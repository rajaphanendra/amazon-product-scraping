import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import NoSuchElementException

options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
options.add_argument("--headless")
options.add_experimental_option("detach", True)

file_path = "Your Excel file Path"
book = openpyxl.load_workbook(file_path)
# sheet = book.active

def get_or_create_sheet(book, sheet_name):
    """Get a sheet by name or create it if it doesn't exist."""
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
    else:
        sheet = book.create_sheet(title=sheet_name)
        sheet.append(["Product Name", "Price", "Rating", "Number of Ratings"])
    return sheet

# Add headers if the sheet is empty
# if sheet.max_row == 1:
#     sheet.append(["Product Name", "Price", "Rating", "Number of Ratings"])

driver = webdriver.Chrome(options=options)
driver.get("https://amazon.com/")
driver.implicitly_wait(4)

# clicking on try different image
try:
    driver.find_element(By.XPATH, "//div[@class='a-row']/div[2]/a").click()
except:
    print("Not detected as bot !!!")

# Search the Item
# search_term = "cricket bat"
search_term = "iphone"
driver.find_element(By.XPATH, "//input[@type='text']").send_keys(search_term)
driver.find_element(By.XPATH, "//input[@type='text']").send_keys(Keys.ENTER)

wait = WebDriverWait(driver, 10)
wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[@data-asin]")))

# Get or create the sheet for the search term
sheet = get_or_create_sheet(book, search_term)

try:
    products = driver.find_elements(By.XPATH, "//div[@data-asin]")
    # print(products)

    for product in products:
        product_asin = product.get_attribute("data-asin")
        print("Product ASIN:", product_asin)
        # product_asin.strip()
        # if len(product_asin) == 10:
        try:
            productName_element = product.find_element(By.XPATH, ".//h2/a/span")
            productName = productName_element.text
            print(productName)
        except NoSuchElementException as e:
            print(f"Skipping product {product_asin} as the name cannot be extracted.")
            continue
                # print(f"Error extracting product name for element {product.get_attribute('outerHTML')}: {e}")

        try:
            product_price_whole_element = product.find_element(By.XPATH, ".//span[@class='a-price-whole']")
            product_price_fraction_element = product.find_element(By.XPATH, ".//span[@class='a-price-fraction']")
            product_price = f"{product_price_whole_element.text}.{product_price_fraction_element.text}"
            print("Product Price:", product_price)
        except NoSuchElementException as e:
            print(f"No price found for product {product_asin}")

            # Extract the product rating
        try:
            wait = WebDriverWait(driver, 5)
            wait.until(EC.presence_of_all_elements_located((By.XPATH, ".//span[contains(@class, 'a-icon-alt')]")))
            product_rating_element = product.find_element(By.XPATH, ".//span[contains(@class, 'a-icon-alt')]")
            # print(product_rating_element)
            product_rating = product_rating_element.get_attribute("innerText").split(' ')[0]
            print("Product Rating:", product_rating)
        except NoSuchElementException as e:
            print(f"No rating found for product {product_asin}")

            # Extract the number of ratings
        try:
            num_ratings_element = product.find_element(By.XPATH, ".//span[contains(@class, 'a-size-base') and contains(@class, 's-underline-text')]")
            num_ratings = num_ratings_element.text
            # num_ratings = num_ratings_text.split()[0]  # Extract the first part which contains the number
            print("Number of Ratings:", num_ratings)
        except NoSuchElementException as e:
            print(f"No number of ratings found for product {product_asin}")

        if len(productName) > 0:
            print("Going to write")
            sheet.append([productName, product_price, product_rating, num_ratings])

    book.save(file_path)

except Exception as e:
    print(f"An error occurred: {e}")
